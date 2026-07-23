import openpyxl
import os
import zipfile
import xml.etree.ElementTree as ET

excel_path = r"SEO report Updated.xlsx"
docx_path = r"Midnight_Stories_Website_Audit_Remediation_Plan.docx"

print("--- EXCEL FILE ANALYSIS ---")
if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n--- Sheet: {name} ({sheet.max_row} rows, {sheet.max_column} cols) ---")
        rows = [r for r in sheet.iter_rows(values_only=True) if any(r)]
        for r in rows[:20]:
            print(r)
else:
    print("Excel file not found!")

print("\n--- WORD DOCX ANALYSIS ---")
if os.path.exists(docx_path):
    with zipfile.ZipFile(docx_path) as z:
        xml_content = z.read("word/document.xml")
        tree = ET.fromstring(xml_content)
        texts = [node.text for node in tree.iter() if node.text]
        full_text = " ".join(texts)
        print("Docx Text Preview (first 1000 chars):")
        print(full_text[:1000])
else:
    print("Docx file not found!")
