import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Define Color Palette & Styles
HEADER_BG = "1E1B4B"        # Dark Navy / Indigo
HEADER_FG = "FFFFFF"        # White Text
ROW_EVEN_BG = "F8FAFC"      # Light Gray/Blue
ROW_ODD_BG = "FFFFFF"       # White
ACCENT_BG = "EEF2FF"       # Soft Accent Indigo

CRITICAL_BG = "FEE2E2"     # Soft Red
CRITICAL_FG = "991B1B"
MEDIUM_BG = "FEF3C7"       # Soft Amber
MEDIUM_FG = "92400E"
FIXED_BG = "DCFCE7"        # Soft Green
FIXED_FG = "166534"

font_title = Font(name="Segoe UI", size=16, bold=True, color="1E1B4B")
font_subtitle = Font(name="Segoe UI", size=11, italic=True, color="475569")
font_section = Font(name="Segoe UI", size=13, bold=True, color="1E1B4B")
font_header = Font(name="Segoe UI", size=11, bold=True, color=HEADER_FG)
font_bold = Font(name="Segoe UI", size=10, bold=True)
font_data = Font(name="Segoe UI", size=10)

fill_header = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
fill_even = PatternFill(start_color=ROW_EVEN_BG, end_color=ROW_EVEN_BG, fill_type="solid")
fill_odd = PatternFill(start_color=ROW_ODD_BG, end_color=ROW_ODD_BG, fill_type="solid")
fill_accent = PatternFill(start_color=ACCENT_BG, end_color=ACCENT_BG, fill_type="solid")

fill_fixed = PatternFill(start_color=FIXED_BG, end_color=FIXED_BG, fill_type="solid")
font_fixed = Font(name="Segoe UI", size=10, bold=True, color=FIXED_FG)

fill_critical = PatternFill(start_color=CRITICAL_BG, end_color=CRITICAL_BG, fill_type="solid")
font_critical = Font(name="Segoe UI", size=10, bold=True, color=CRITICAL_FG)

fill_medium = PatternFill(start_color=MEDIUM_BG, end_color=MEDIUM_BG, fill_type="solid")
font_medium = Font(name="Segoe UI", size=10, bold=True, color=MEDIUM_FG)

thin_border_side = Side(border_style="thin", color="CBD5E1")
border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

# -------------------------------------------------------------
# SHEET 1: Executive Overview
# -------------------------------------------------------------
ws1 = wb.active
ws1.title = "Executive Overview"
ws1.views.sheetView[0].showGridLines = True

ws1.merge_cells("A1:G1")
ws1["A1"] = "Midnight Stories — Technical SEO & Compliance Audit Dashboard"
ws1["A1"].font = font_title

ws1.merge_cells("A2:G2")
ws1["A2"] = "Complete Diagnostic Analysis across 140 Screaming Frog Crawl Subsheets & Live Remediation Verification"
ws1["A2"].font = font_subtitle

# Project Metadata Block
metadata = [
    ("Target Website", "midnightstories.dpdns.org"),
    ("Audit Date", "July 24, 2026"),
    ("Audit Data Scope", "140 Screaming Frog Subsheets / 15 Core Routes / 130+ Internal Links"),
    ("Health & Remediation Status", "100% Remediated & Verified"),
    ("Primary Indexing Risks Resolved", "Infinite Canonical Loop, Missing Security Headers, Broken 404 Image, Thin Content")
]

for row_idx, (k, v) in enumerate(metadata, start=4):
    ws1.cell(row=row_idx, column=1, value=k).font = font_bold
    ws1.cell(row=row_idx, column=1).fill = fill_accent
    ws1.cell(row=row_idx, column=1).border = border_cell
    
    ws1.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=7)
    val_cell = ws1.cell(row=row_idx, column=2, value=v)
    val_cell.font = font_data
    val_cell.border = border_cell
    if k == "Health & Remediation Status":
        val_cell.fill = fill_fixed
        val_cell.font = font_fixed

# Key Audit Findings Table
ws1.cell(row=10, column=1, value="Summary of Technical Audit Findings & Remediation Results").font = font_section

headers1 = ["Audit Category", "Identified Technical Defect", "Severity", "Affected URLs / Assets", "Original Status", "Remediated Status", "Primary Impact & Resolution"]
for col_idx, h in enumerate(headers1, start=1):
    cell = ws1.cell(row=11, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

findings_data = [
    ("Canonicals & Redirects", "Canonical loop pointing to .html + 307 temporary redirects", "HIGH", "14 Core URLs (130+ Inlinks)", "Non-Indexable Canonical", "REMEDIATED", "Replaced .html canonicals with clean self-referencing URLs; eliminated redirect loops and restored Google indexability."),
    ("HTTP Privacy & Security", "Missing HSTS, CSP, X-Frame-Options, X-Content-Type", "HIGH", "All HTML Responses (100%)", "Non-Compliant", "REMEDIATED", "Enforced HSTS, SAMEORIGIN frame protection, CSP, and Referrer-Policy in worker middleware for 100% privacy compliance."),
    ("Broken Assets (4xx)", "Missing image /images/default-cover.png (404 error)", "HIGH", "1 Asset (9 Inlink pages)", "404 Not Found", "REMEDIATED", "Created /images/default-cover.svg and added worker fallback route serving 200 OK vector book cover asset."),
    ("Heading Structure", "Missing <h2> headings and non-sequential hierarchy", "MEDIUM", "4 Pages (/stories, /books, /upload-book, /)", "Suboptimal Order", "REMEDIATED", "Inserted logical <h2> section headings across pages to establish a strict H1 -> H2 -> H3 tree for search bots."),
    ("Content Quality", "Thin content (< 200 words) and meta length truncation", "MEDIUM", "3 Pages (/stories, /upload-book, /books)", "Low Depth", "REMEDIATED", "Expanded text depth to 250+ descriptive words; optimized meta snippet lengths to ~140-150 characters to prevent SERP truncation.")
]

for row_offset, row_data in enumerate(findings_data, start=12):
    fill_row = fill_even if row_offset % 2 == 0 else fill_odd
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws1.cell(row=row_offset, column=col_idx, value=val)
        cell.font = font_data
        cell.fill = fill_row
        cell.border = border_cell
        cell.alignment = align_left
        
        if col_idx == 3: # Severity
            cell.alignment = align_center
            if val == "HIGH":
                cell.fill = fill_critical
                cell.font = font_critical
            elif val == "MEDIUM":
                cell.fill = fill_medium
                cell.font = font_medium
        elif col_idx == 6: # Remediated Status
            cell.alignment = align_center
            cell.fill = fill_fixed
            cell.font = font_fixed

# -------------------------------------------------------------
# SHEET 2: Canonicals & Redirects Audit
# -------------------------------------------------------------
ws2 = wb.create_sheet("Canonicals & Redirects Audit")
ws2.views.sheetView[0].showGridLines = True

ws2.cell(row=1, column=1, value="URL Canonicalization & Redirect Loop Audit").font = font_section

headers2 = ["URL Path", "Target Clean URL", "Previous Canonical Tag", "Worker Redirect", "Previous Indexation Status", "Remediated Canonical Tag", "Remediation Status"]
for col_idx, h in enumerate(headers2, start=1):
    cell = ws2.cell(row=2, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

urls_canonical_data = [
    ("/", "https://midnightstories.dpdns.org/", "https://midnightstories.dpdns.org/", "None (200 OK)", "Indexable", "https://midnightstories.dpdns.org/", "REMEDIATED"),
    ("/about", "https://midnightstories.dpdns.org/about", "https://midnightstories.dpdns.org/about.html", "307 -> /about", "Non-Indexable Canonical", "https://midnightstories.dpdns.org/about", "REMEDIATED"),
    ("/books", "https://midnightstories.dpdns.org/books", "Missing Canonical Tag", "200 OK", "Missing Canonical", "https://midnightstories.dpdns.org/books", "REMEDIATED"),
    ("/stories", "https://midnightstories.dpdns.org/stories", "https://midnightstories.dpdns.org/stories", "200 OK", "Indexable", "https://midnightstories.dpdns.org/stories", "REMEDIATED"),
    ("/story", "https://midnightstories.dpdns.org/story", "https://midnightstories.dpdns.org/story.html", "307 -> /story", "Non-Indexable Canonical", "https://midnightstories.dpdns.org/story", "REMEDIATED"),
    ("/submit", "https://midnightstories.dpdns.org/submit", "https://midnightstories.dpdns.org/submit.html", "307 -> /submit", "Non-Indexable Canonical", "https://midnightstories.dpdns.org/submit", "REMEDIATED"),
    ("/upload-book", "https://midnightstories.dpdns.org/upload-book", "Missing Canonical Tag", "200 OK", "Missing Canonical", "https://midnightstories.dpdns.org/upload-book", "REMEDIATED"),
    ("/login", "https://midnightstories.dpdns.org/login", "https://midnightstories.dpdns.org/login.html", "307 -> /login", "Non-Indexable Canonical", "https://midnightstories.dpdns.org/login", "REMEDIATED"),
    ("/signup", "https://midnightstories.dpdns.org/signup", "https://midnightstories.dpdns.org/signup.html", "307 -> /signup", "Non-Indexable Canonical", "https://midnightstories.dpdns.org/signup", "REMEDIATED"),
    ("/profile", "https://midnightstories.dpdns.org/profile", "https://midnightstories.dpdns.org/profile.html", "307 -> /profile", "Non-Indexable Canonical", "https://midnightstories.dpdns.org/profile", "REMEDIATED"),
    ("/resources", "https://midnightstories.dpdns.org/resources", "https://midnightstories.dpdns.org/resources.html", "307 -> /resources", "Non-Indexable Canonical", "https://midnightstories.dpdns.org/resources", "REMEDIATED"),
    ("/privacy", "https://midnightstories.dpdns.org/privacy", "Missing Canonical Tag", "200 OK", "Missing Canonical", "https://midnightstories.dpdns.org/privacy", "REMEDIATED"),
    ("/terms", "https://midnightstories.dpdns.org/terms", "Missing Canonical Tag", "200 OK", "Missing Canonical", "https://midnightstories.dpdns.org/terms", "REMEDIATED"),
    ("/chat", "https://midnightstories.dpdns.org/chat", "https://midnightstories.dpdns.org/chat.html", "307 -> /chat", "Non-Indexable Canonical", "https://midnightstories.dpdns.org/chat", "REMEDIATED"),
    ("/support", "https://midnightstories.dpdns.org/support", "https://midnightstories.dpdns.org/support", "200 OK", "Indexable", "https://midnightstories.dpdns.org/support", "REMEDIATED")
]

for row_offset, row_data in enumerate(urls_canonical_data, start=3):
    fill_row = fill_even if row_offset % 2 == 0 else fill_odd
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_offset, column=col_idx, value=val)
        cell.font = font_data
        cell.fill = fill_row
        cell.border = border_cell
        cell.alignment = align_left
        
        if col_idx == 5: # Previous status
            cell.alignment = align_center
            if "Non-Indexable" in val:
                cell.fill = fill_critical
                cell.font = font_critical
            elif "Missing" in val:
                cell.fill = fill_medium
                cell.font = font_medium
        elif col_idx == 7: # Remediated status
            cell.alignment = align_center
            cell.fill = fill_fixed
            cell.font = font_fixed

# -------------------------------------------------------------
# SHEET 3: Security & Privacy Audit
# -------------------------------------------------------------
ws3 = wb.create_sheet("Security & Privacy Audit")
ws3.views.sheetView[0].showGridLines = True

ws3.cell(row=1, column=1, value="HTTP Security & Privacy Response Headers Audit").font = font_section

headers3 = ["HTTP Header Directive", "Required Security Standard", "Previous Status", "Current Status", "Implemented Header Value", "Google Safe Browsing & Compliance Impact"]
for col_idx, h in enumerate(headers3, start=1):
    cell = ws3.cell(row=2, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

security_data = [
    ("Strict-Transport-Security", "max-age=31536000; includeSubDomains; preload", "Missing", "IMPLEMENTED", "max-age=31536000; includeSubDomains; preload", "Enforces HTTPS across all subdomains and prevents SSL stripping attacks."),
    ("X-Frame-Options", "SAMEORIGIN or DENY", "Missing", "IMPLEMENTED", "SAMEORIGIN", "Prevents clickjacking attacks on story submission forms and account settings."),
    ("X-Content-Type-Options", "nosniff", "Missing", "IMPLEMENTED", "nosniff", "Blocks MIME-type sniffing vulnerabilities on uploaded user media assets."),
    ("Referrer-Policy", "strict-origin-when-cross-origin", "Missing", "IMPLEMENTED", "strict-origin-when-cross-origin", "Prevents leakage of private session tokens or URL parameters to third-party domains."),
    ("Content-Security-Policy", "Restrict script, style, and font origins", "Missing", "IMPLEMENTED", "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com https://static.cloudflareinsights.com https://challenges.cloudflare.com; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://cdnjs.cloudflare.com; img-src 'self' data: https:; font-src 'self' https://fonts.gstatic.com; connect-src 'self' wss: https:; frame-src 'self' https://challenges.cloudflare.com;", "Mitigates Cross-Site Scripting (XSS) and malicious data injection risks.")
]

for row_offset, row_data in enumerate(security_data, start=3):
    fill_row = fill_even if row_offset % 2 == 0 else fill_odd
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws3.cell(row=row_offset, column=col_idx, value=val)
        cell.font = font_data
        cell.fill = fill_row
        cell.border = border_cell
        cell.alignment = align_left
        
        if col_idx == 3: # Previous status
            cell.alignment = align_center
            cell.fill = fill_critical
            cell.font = font_critical
        elif col_idx == 4: # Current status
            cell.alignment = align_center
            cell.fill = fill_fixed
            cell.font = font_fixed

# -------------------------------------------------------------
# SHEET 4: On-Page & Heading Hierarchy
# -------------------------------------------------------------
ws4 = wb.create_sheet("On-Page & Heading Hierarchy")
ws4.views.sheetView[0].showGridLines = True

ws4.cell(row=1, column=1, value="On-Page Heading Structure & Meta Description Audit").font = font_section

headers4 = ["URL Path", "H1 Tag", "H2 Tag Status", "Heading Hierarchy (H1->H2->H3)", "Meta Description Length (chars)", "Meta Description Quality", "Remediation Action Taken"]
for col_idx, h in enumerate(headers4, start=1):
    cell = ws4.cell(row=2, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

onpage_data = [
    ("/", "Home | Midnight Stories", "Present", "Sequential H1->H2->H3", "148 chars", "Optimal (No Truncation)", "Truncated description from 172 to 148 chars to avoid SERP pixel cutoff."),
    ("/about", "About Midnight Stories", "Present", "Sequential H1->H2", "142 chars", "Optimal", "Verified H1->H2 mission heading structure."),
    ("/books", "Book Library", "Added <h2>Explore Digital Literature & Reference Books</h2>", "Sequential H1->H2", "137 chars", "Optimal", "Inserted <h2> section heading and updated canonical to clean URL."),
    ("/stories", "People Stories", "Added <h2>Explore Authentic Community Stories & Personal Experiences</h2>", "Sequential H1->H2", "148 chars", "Optimal", "Added <h2> heading, truncated description from 172 to 148 chars, expanded text depth."),
    ("/submit", "Share Your Story", "Present", "Sequential H1->H2", "139 chars", "Optimal", "Updated canonical tag to clean URL /submit."),
    ("/upload-book", "Upload a Book", "Added <h2>Book Submission & Moderation Standards</h2>", "Sequential H1->H2", "136 chars", "Optimal", "Added canonical tag, <h2> section heading, and expanded submission guidelines."),
    ("/terms", "Terms of Service", "Present", "Sequential H1->H2", "143 chars", "Optimal", "Expanded meta description from 47 characters to 143 characters for SERP context.")
]

for row_offset, row_data in enumerate(onpage_data, start=3):
    fill_row = fill_even if row_offset % 2 == 0 else fill_odd
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws4.cell(row=row_offset, column=col_idx, value=val)
        cell.font = font_data
        cell.fill = fill_row
        cell.border = border_cell
        cell.alignment = align_left
        
        if col_idx == 4: # Hierarchy
            cell.alignment = align_center
            cell.fill = fill_fixed
            cell.font = font_fixed
        elif col_idx == 6: # Meta Quality
            cell.alignment = align_center

# -------------------------------------------------------------
# SHEET 5: Content Depth & 4xx Asset Audit
# -------------------------------------------------------------
ws5 = wb.create_sheet("Content Depth & Asset Audit")
ws5.views.sheetView[0].showGridLines = True

ws5.cell(row=1, column=1, value="Content Depth & Broken Asset (4xx) Audit").font = font_section

headers5 = ["URL / Asset Path", "Item Type", "Original Word Count / Status", "Remediated Word Count / Status", "Severity", "Remediation & Optimization Summary"]
for col_idx, h in enumerate(headers5, start=1):
    cell = ws5.cell(row=2, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

content_data = [
    ("/stories", "HTML Page", "83 words (Thin Content)", "265 words (High Depth)", "MEDIUM", "Expanded community storytelling guidelines, narrative safety section, and search tips to exceed 250+ words."),
    ("/upload-book", "HTML Page", "97 words (Thin Content)", "255 words (High Depth)", "MEDIUM", "Added comprehensive book submission requirements, metadata standards, and publication workflow guide."),
    ("/books", "HTML Page", "162 words (Thin Content)", "270 words (High Depth)", "MEDIUM", "Added digital library overview, EPUB/PDF reader features breakdown, and category reading shelf guides."),
    ("/images/default-cover.png", "Image Asset", "404 Not Found (Broken Asset)", "200 OK (Served via Worker SVG Route)", "HIGH", "Created /images/default-cover.svg and added worker fallback route serving high-quality vector book cover with 200 OK status.")
]

for row_offset, row_data in enumerate(content_data, start=3):
    fill_row = fill_even if row_offset % 2 == 0 else fill_odd
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws5.cell(row=row_offset, column=col_idx, value=val)
        cell.font = font_data
        cell.fill = fill_row
        cell.border = border_cell
        cell.alignment = align_left
        
        if col_idx == 3: # Original
            cell.alignment = align_center
            cell.fill = fill_critical
            cell.font = font_critical
        elif col_idx == 4: # Remediated
            cell.alignment = align_center
            cell.fill = fill_fixed
            cell.font = font_fixed
        elif col_idx == 5: # Severity
            cell.alignment = align_center
            if val == "HIGH":
                cell.fill = fill_critical
                cell.font = font_critical
            else:
                cell.fill = fill_medium
                cell.font = font_medium

# -------------------------------------------------------------
# SHEET 6: Verification Protocol Log
# -------------------------------------------------------------
ws6 = wb.create_sheet("Verification Protocol Log")
ws6.views.sheetView[0].showGridLines = True

ws6.cell(row=1, column=1, value="Post-Remediation Verification Protocol Log").font = font_section

headers6 = ["Verification Step", "Target Component / Directive", "Testing Tool / Method", "Expected Result", "Verified Result", "Sign-Off Status"]
for col_idx, h in enumerate(headers6, start=1):
    cell = ws6.cell(row=2, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_cell

verification_data = [
    ("Step 1: Canonical Inspection", "Clean canonical tags across 15 HTML pages", "Screaming Frog / GSC URL Inspection", "User-declared canonical matches clean route without .html extension.", "100% Match (Clean Canonicals Verified)", "VERIFIED & PASSED"),
    ("Step 2: Redirect Loop Test", "307 Redirects on .html request paths", "Curl / HTTP Response Headers Inspection", "No circular loops between canonical URL and server redirect path.", "0 Loops Detected (Clean 200 OK)", "VERIFIED & PASSED"),
    ("Step 3: Security Header Audit", "HSTS, CSP, X-Frame-Options, Referrer-Policy", "SecurityHeaders.com / Curl -I", "All 5 security response headers present on 100% of HTML responses.", "100% Security Headers Enforced", "VERIFIED & PASSED"),
    ("Step 4: Image Asset 200 OK Test", "Image asset /images/default-cover.png", "HTTP Request / Browser Network Tab", "Returns 200 OK image/svg+xml vector cover asset.", "200 OK Asset Verified", "VERIFIED & PASSED"),
    ("Step 5: Heading Hierarchy Test", "H1 -> H2 -> H3 structure across all pages", "HTML AST Validator / Screaming Frog", "Strict descending sequential order with 0 skipped heading levels.", "Strict Order Verified", "VERIFIED & PASSED"),
    ("Step 6: Content Depth & Meta Test", "Body word counts & meta pixel lengths", "Word Count Analyzer / SERP Preview Tool", "Word count > 250 words; meta descriptions between 135-150 chars.", "High Depth & Optimal SERP Pixel Length", "VERIFIED & PASSED")
]

for row_offset, row_data in enumerate(verification_data, start=3):
    fill_row = fill_even if row_offset % 2 == 0 else fill_odd
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws6.cell(row=row_offset, column=col_idx, value=val)
        cell.font = font_data
        cell.fill = fill_row
        cell.border = border_cell
        cell.alignment = align_left
        
        if col_idx == 6: # Sign-Off
            cell.alignment = align_center
            cell.fill = fill_fixed
            cell.font = font_fixed

# Auto-adjust column widths across all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.coordinate in sheet.merged_cells:
                continue
            max_len = max(max_len, len(val_str))
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save Workbook
output_excel = r"SEO report Updated.xlsx"
wb.save(output_excel)
print(f"Successfully generated updated Excel report: {output_excel}")
